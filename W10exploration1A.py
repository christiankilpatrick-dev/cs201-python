#W10 Exploration 1 A

import openpyxl

def lookup_superfund_sites(desired_state):
    """
    Returns a list containing information about superfund sites in the state.
    Provide a two-character state abbreviation, such as TX
    """
    if len(desired_state) != 2:
        raise Exception('Please enter a two-character state abbreviation, such as OR')

    SUPERFUND_XLSX = "superfund_npl.xlsx"
    wb = openpyxl.load_workbook(SUPERFUND_XLSX)
    sheet = wb.active

    # Keep track of which column corresponds to each header, numbered from 0.
    # Note that cells in .cell() are numbered from 1 (not 0).
    # It's your job to pay attention to such things!
    header_to_colnum = {}
    for c in range(1, sheet.max_column): # through max column
        value = str(sheet.cell(1, c).value).strip()
        header_to_colnum[value] = c - 1

    col_of_name = header_to_colnum.get('Name', 0)
    col_of_address = header_to_colnum.get('Address', 0)
    col_of_city = header_to_colnum.get('City', 0)
    col_of_state = header_to_colnum.get('State', 0)

    if col_of_name == 0 or col_of_address == 0 or col_of_city == 0 or col_of_state == 0:
        raise Exception('Unable to find headers: Name, Address, City, State')

    locations = []
    for row in sheet.max_row:
        # Check of state matches our desired.
        state = str(row[col_of_state].value).strip()
        if state == desired_state:
            name = str(row[col_of_name].xxxx).xxxx()
            address = str(row[col_of_address].xxxx).xxxx()
            city = str(row[col_of_city].xxxx).xxxx()
            locations.append(f'{name}, {address}, {city} {state}')
    return locations

print("Where do you live?")
home_state = input()
superfund_sites = lookup_superfund_sites(home_state)
superfund_sites = sorted(superfund_sites)

for site in superfund_sites:
    print(site)
